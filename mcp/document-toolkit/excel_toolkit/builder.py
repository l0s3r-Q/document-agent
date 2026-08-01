"""builder.py —— 按 ExcelSpec 生成 .xlsx：数据、样式美化、合并、列宽、筛选。"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")  # 浅蓝
_HEADER_FONT = Font(bold=True, size=11)
_BODY_FONT = Font(size=11)

# 常用填充色（供美化参考）
FILL_COLORS = {
    "blue": "D9E2F3", "green": "E2EFDA", "orange": "FCE4D6",
    "yellow": "FFF2CC", "grey": "EDEDED", "red": "F8CBAD",
}


def _apply_cell_style(cell, style: dict | None):
    """单元格样式：font{name,size,bold,color}、align{horizontal,vertical,wrap}、fill、border。"""
    s = style or {}
    font = s.get("font") or {}
    cell.font = Font(
        name=font.get("name", "宋体"),
        size=font.get("size", 11),
        bold=font.get("bold", False),
        color=font.get("color", "000000"),
    )
    align = s.get("align") or {}
    cell.alignment = Alignment(
        horizontal=align.get("horizontal", "center"),
        vertical=align.get("vertical", "center"),
        wrap_text=align.get("wrap", True),
    )
    fill = s.get("fill")
    if fill:
        color = FILL_COLORS.get(fill, fill)  # 支持名称或 hex
        cell.fill = PatternFill("solid", fgColor=str(color).lstrip("#"))
    if s.get("border", True):
        cell.border = _BORDER


def build(spec: dict, output_path: str) -> dict:
    """按 ExcelSpec 生成 xlsx。

    spec 契约:
      {sheets: [
        {name?, rows: [[...]], styles?{header?, body?}, merges?[{range}],
         col_widths?{col_letter: width}, freeze?("A2"), filter?(bool),
         header_row?(bool, 默认首行视为表头), fill?("blue"|hex)}
      ]}
    """
    if not output_path.lower().endswith(".xlsx"):
        return {"ok": False, "error": "输出路径必须以 .xlsx 结尾"}
    sheets = spec.get("sheets") or []
    if not sheets:
        return {"ok": False, "error": "sheets 不能为空"}

    wb = Workbook()
    wb.remove(wb.active)  # 移除默认 sheet，按 spec 创建

    for si, sh in enumerate(sheets):
        rows = sh.get("rows") or []
        if not rows:
            continue
        name = sh.get("name") or f"Sheet{si + 1}"
        ws = wb.create_sheet(title=name[:31])

        header_row = sh.get("header_row", True)
        header_style = sh.get("styles", {}).get("header") or {
            "font": {"bold": True}, "fill": sh.get("fill", "blue")}
        body_style = sh.get("styles", {}).get("body") or {}

        for ri, row in enumerate(rows, 1):
            for ci, value in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=value)
                if header_row and ri == 1:
                    _apply_cell_style(cell, header_style)
                else:
                    _apply_cell_style(cell, body_style)

        # 合并单元格
        for m in sh.get("merges") or []:
            ws.merge_cells(m["range"])

        # 列数基准：取所有行的最大列数（防空首行崩溃）
        ncols = max((len(r) for r in rows), default=0)

        # 列宽：显式设置或按内容自适应（L1：中英文分计宽度）
        col_widths = sh.get("col_widths") or {}
        for ci in range(1, ncols + 1):
            letter = get_column_letter(ci)
            if letter in col_widths:
                ws.column_dimensions[letter].width = col_widths[letter]
            else:
                def _width(v):
                    return sum(2 if ord(c) > 127 else 1 for c in str(v))
                max_len = max((_width(r[ci - 1]) for r in rows if ci - 1 < len(r)), default=6)
                ws.column_dimensions[letter].width = min(max(max_len + 4, 10), 50)

        # 冻结窗格 + 自动筛选（freeze 格式预校验：如 "A2"）
        import re as _re
        _freeze = sh.get("freeze")
        if _freeze and _re.fullmatch(r"[A-Za-z]{1,3}[1-9][0-9]{0,6}", str(_freeze)):
            ws.freeze_panes = _freeze
        elif header_row and len(rows) > 1:
            ws.freeze_panes = "A2"
        if sh.get("filter", True) and header_row and len(rows) > 1 and ncols > 0:
            ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{len(rows)}"

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return {"ok": True, "path": output_path, "sheets": len(wb.sheetnames), "sheet_names": wb.sheetnames}
