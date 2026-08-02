"""parser.py —— 读取 .xlsx：sheet/行数据/合并/列宽。"""

from __future__ import annotations

from openpyxl import load_workbook


def parse(path: str, sheet_name: str | None = None) -> dict:
    """解析 xlsx 文件。

    sheet_name 为空时解析全部 sheet；返回 rows/merged/col_widths。
    """
    wb = load_workbook(path, data_only=False)
    # 公式缓存值需 data_only=True 的二次加载（数据源场景依赖缓存值）
    try:
        wb_cached = load_workbook(path, data_only=True)
    except Exception:  # noqa: BLE001
        wb_cached = None
    result = {"file": path, "sheets": []}

    for ws in wb.worksheets:
        if sheet_name and ws.title != sheet_name:
            continue
        cached_ws = wb_cached[ws.title] if wb_cached is not None and ws.title in wb_cached.sheetnames else None
        rows = []
        for ri, row in enumerate(ws.iter_rows()):
            vals = []
            for ci, cell in enumerate(row):
                if cell.data_type == "f":  # 公式：标注公式与缓存值
                    cached = None
                    if cached_ws is not None:
                        try:
                            cached = cached_ws.cell(row=ri + 1, column=ci + 1).value
                        except Exception:  # noqa: BLE001
                            cached = None
                    vals.append({"formula": cell.value, "cached": cached})
                else:
                    vals.append("" if cell.value is None else cell.value)
            rows.append(vals)
        merged = [str(m) for m in ws.merged_cells.ranges]
        col_widths = {letter: dim.width for letter, dim in ws.column_dimensions.items() if dim.width}
        result["sheets"].append({
            "name": ws.title,
            "rows": rows,
            "merged": merged,
            "col_widths": col_widths,
        })
        if sheet_name:
            break
    return result


def to_data(path: str, sheet_name: str | None = None) -> dict:
    """把 Excel 数据表转为 JSON 行数组（首行作为字段名），供批量生成等使用。

    返回 {ok, rows: [{字段: 值, ...}, ...], sheet, skipped_rows}
    """
    data = parse(path, sheet_name)
    if not data["sheets"]:
        return {"ok": False, "error": "未找到 sheet"}
    ws = data["sheets"][0]
    rows = ws["rows"]
    if not rows:
        return {"ok": False, "error": "sheet 为空"}
    headers = []
    seen = {}
    for h in rows[0]:
        hs = str(h).strip() if h is not None else ""
        if hs:
            n = seen.get(hs, 0)
            seen[hs] = n + 1
            headers.append(hs if n == 0 else f"{hs}_{n + 1}")
        else:
            headers.append("")
    out_rows = []
    skipped = 0
    for r in rows[1:]:
        if all(str(v).strip() == "" for v in r):
            skipped += 1
            continue
        item = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            v = r[i] if i < len(r) else ""
            if isinstance(v, dict):  # 公式单元格：用缓存值（无缓存则公式串）
                v = v.get("cached")
                if v is None:
                    v = v if False else v.get("formula", "")
                    v = "" if v is None else v
            item[h] = "" if v is None else v
        out_rows.append(item)
    return {"ok": True, "rows": out_rows, "sheet": ws["name"], "skipped_rows": skipped}
